########################################################################################
# エクセルファイル操縦クラス
#[概要]ワークブックの操縦を行います
########################################################################################

import openpyxl
import os
import PySimpleGUI
from openpyxl.styles.borders import Border, Side

class operate_excel:
    
    # テンプレートエクセルファイル利用
    use_template_flag = None
    # 出力フォルダ名（フルパス）
    output_absolute_dir_name = None
    # 出力ファイル名（フルパス）
    output_absolute_file_name = None
    # データ登録シート名
    data_sheet_name = None
    # ワークブックオブジェクト
    workbook = None
    # ワークシートオブジェクト
    worksheet = None
    # データ書き込み行番号
    target_row = 1
    # データ書き込み列番号
    target_col = 0
    
    # 罫線スタイル
    side = Side(style='thin', color='000000')
    # 罫線設定
    border = Border(top=side, bottom=side, left=side, right=side)
    # 背景色
    fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='BDD7EE', bgColor='BDD7EE')
    
    
    ########################################################################################
    # コンストラクタ
    # [概要]出力するワークブックオブジェクトを生成し、ワークブックオブジェクトに対してデータを登録するシートを追加します。
    # [引数1]ツールルートディレクトリパス
    # [引数2]ツール設定辞書オブジェクト
    ########################################################################################
    def __init__(self, root_dir_path, setting_dict):
        
        self.output_absolute_dir_name = os.path.join(root_dir_path, setting_dict['output_dir'])
        self.output_absolute_file_name = os.path.join(root_dir_path, setting_dict['output_dir'], setting_dict['output_file_name'])
        
        # テンプレートファイルを使用する場合はテンプレートファイルを読み込んでオブジェクトを生成する
        if setting_dict['use_template'] == "yes":
            self.workbook = openpyxl.load_workbook(os.path.join(root_dir_path,setting_dict['template_dir'],setting_dict['template_file_name']))
            self.use_template_flag = True
        else:
            self.workbook = openpyxl.Workbook()
            self.use_template_flag = False
        
        # PySimpleGUIでシート名を受け取る
        # PySimpleGUIウィンドウ設定
        PySimpleGUI.theme('Dark Blue 3')
        # 表示内容設定
        layout = [
            [PySimpleGUI.Text('データを登録するシート名を入力してください。')],
            [PySimpleGUI.Text('既存のシート名を入力するとシートを上書きします。')],
            [PySimpleGUI.Text('シート名', size=(31, 1)), PySimpleGUI.InputText('')],
            [PySimpleGUI.Submit(button_text='設定')]
        ]

        # PySimpleGUIウィンドウ表示
        window = PySimpleGUI.Window('', layout)
        
        while True:
            event, values = window.read()

            if event is None:
                break
            if event == '設定':
                self.data_sheet_name = values[0]
                break
                
        # PySimpleGUIウィンドウの破棄と終了
        window.close()
        
        # 入力されたシートをワークブックに加えるが、既存のシートの場合は事前に削除する。
        if (self.data_sheet_name in self.workbook.sheetnames):
            self.workbook.remove_sheet(self.data_sheet_name)
        self.workbook.create_sheet(self.data_sheet_name, len(self.workbook.sheetnames))
        
        # 書き込むワークシートを設定
        self.worksheet = self.workbook[self.data_sheet_name]    
        

    ########################################################################################
    # テーブル物理名/論理名書き込み処理
    # [概要]引数のテーブル物理名/論理名を格納したデータフレームの内容をエクセルに書き込みます。
    # [引数1]テーブル物理名/論理名データフレーム
    ########################################################################################
    def write_table_names(self, table_names_df):
        # セルのデコレーション
        self.decorate_cell(self.get_target_row_and_increment(True), 1, True)
        self.decorate_cell(self.get_target_row_and_increment(), 2)
        # 値設定
        self.worksheet.cell(row = self.get_target_row_and_increment(), column = 1).value = "テーブル物理名"
        self.worksheet.cell(row = self.get_target_row_and_increment(), column = 2).value = table_names_df.iloc[0, 0]
        # セルのデコレーション
        self.decorate_cell(self.get_target_row_and_increment(True), 1, True)
        self.decorate_cell(self.get_target_row_and_increment(), 2)
        # 値設定
        self.worksheet.cell(row = self.get_target_row_and_increment(), column = 1).value = "テーブル論理名"
        self.worksheet.cell(row = self.get_target_row_and_increment(), column = 2).value = table_names_df.iloc[0, 1]
        
        
    ########################################################################################
    # カラム物理名/論理名書き込み処理
    # [概要]引数のカラム物理名/論理名を格納したデータフレームの内容をエクセルに書き込みます。（カラム名の前に検索条件を書き出します）
    # [引数1]カラム物理名/論理名データフレーム
    # [引数2]クエリ条件辞書
    ########################################################################################
    def write_column_names(self ,column_names_df ,query_condition_dict):
        # セルのデコレーション
        self.decorate_cell(self.get_target_row_and_increment(True), 1, True)
        self.decorate_cell(self.get_target_row_and_increment(), 2)
        
        # カラム名の書き出し前に検索条件を書き出す
        self.worksheet.cell(row = self.get_target_row_and_increment(), column = 1).value = "検索条件"
        self.worksheet.cell(row = self.get_target_row_and_increment(), column = 2).value = query_condition_dict["where"]
        
        # カラム名の書き出し前に書き込む行数をインクリメントしておく
        self.get_target_row_and_increment(True)
        
        # カラム名をループで書き出す
        for index, row in column_names_df.iterrows():
            # セルのデコレーション
            self.decorate_cell(self.get_target_row_and_increment(), index + 1, True)
            self.decorate_cell(self.get_target_row_and_increment() + 1, index + 1, True)
            # カラムの論理名/物理名設定
            self.worksheet.cell(row = self.get_target_row_and_increment(), column = index + 1).value = row[2]
            self.worksheet.cell(row = self.get_target_row_and_increment() + 1, column = index + 1).value = row[1]
    
        # カラム名の書き出し後に書き込む行数をインクリメントしておく
        #（カラム名は論物名を書き込むために2行使うためインクリメントしないと値を書き込む時に整合性が取れなくなる）
        self.get_target_row_and_increment(True)
        
        
    ########################################################################################
    # 取得データ書き込み処理
    # [概要]引数の取得データを格納したデータフレームの内容をエクセルに書き込みます。
    # [引数1]取得データデータフレーム
    ########################################################################################
    def write_datas(self, datas_df):
        # データフレームを1行ずつ抜き出す
        for row_index, row in datas_df.iterrows():
            
            # 書き込み先の行番号をインクリメント
            self.get_target_row_and_increment(True)
            
            # 各列の値を書き込み
            col_index = 0
            for value in row:
                # セルのデコレーション
                self.decorate_cell(self.get_target_row_and_increment(), col_index + 1)
                # 値書き込み
                self.worksheet.cell(row = self.get_target_row_and_increment(), column = col_index + 1).value = value
                col_index = col_index + 1
                
        # 書き込み先の行番号をインクリメント
        self.get_target_row_and_increment(True)
            
        
    ########################################################################################
    # 書き込み行番号取得およびインクリメント処理
    # [概要]書き込み先の行番号を取得します。（引数がTrueであればクラス変数の「データ書き込み行番号」をインクリメントします。）
    # [引数1]インクリメントフラグ（default=False）
    ########################################################################################
    def get_target_row_and_increment(self, increment_flag = False):
        if increment_flag:
            self.target_row = self.target_row + 1
        return self.target_row
    
    ########################################################################################
    # セル装飾処理
    # [概要]引数の行番、列番のセルを枠付き、文字列に装飾します。タイトルカラムであれば背景も設定します。
    # [引数1]行番号
    # [引数2]列番号
    # [引数3]タイトルフラグ
    ########################################################################################
    def decorate_cell(self, row_number, col_number, title_flag = False):
        target_cell = self.worksheet.cell(row = row_number, column = col_number)
        # 罫線設定
        target_cell.border = self.border
        # 表示形式
        target_cell.number_format  = openpyxl.styles.numbers.FORMAT_TEXT
        # タイトルカラム用設定
        if title_flag:
            target_cell.fill = self.fill
        
    ########################################################################################
    # エクセルファイル保存処理
    # [概要]エクセルファイルを保存します。
    # [引数1]ツールルートパス
    # [引数2]ツール設定辞書
    ########################################################################################
    def save_excel_file(self):
        # フォルダを作成する
        os.makedirs(self.output_absolute_dir_name, exist_ok=True)
        
        # 新規ワークブックを作成した場合、不要なシートができてしまうため削除
        if self.use_template_flag == False:
            self.workbook.remove(self.workbook['Sheet'])
            
        self.workbook.save(self.output_absolute_file_name)